"""제안목차 생성 + 엑셀 변환 서비스"""
import json
import logging
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.prompts.outline_generation import build_outline_prompt_dynamic
from backend.services.llm import LLMError, LLMRequest, call_with_fallback
from backend.services.llm_config_store import get_active_config
from backend.services.prompt_store import get_prompt
from backend.services import progress_store
from backend.services.outline_types import OUTLINE_TYPES, get_label, is_supported

logger = logging.getLogger(__name__)


class UnsupportedProjectTypeError(Exception):
    """ISP/ISMP 등 지원되지 않는 사업 유형."""


def _parse_json_response(text: str) -> dict:
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        json_lines: list[str] = []
        inside = False
        for line in lines:
            if line.strip().startswith("```") and not inside:
                inside = True
                continue
            if line.strip() == "```" and inside:
                break
            if inside:
                json_lines.append(line)
        text = "\n".join(json_lines)
    return json.loads(text)


# ---------------------------------------------------------------------------
# 생성
# ---------------------------------------------------------------------------


async def generate_outline(notice_id: int, db) -> dict:
    """대상 공고의 제안목차 생성 (koneps 스키마).

    - analysis_results에서 완료된 분석 + project_type 로드
    - Gemini 호출로 목차 트리 생성
    - proposal_outlines.sections(JSONB) UPSERT (notice_id 기준 1건)

    UnsupportedProjectTypeError 또는 다른 예외 raise.
    """
    from datetime import datetime
    from sqlalchemy import select
    from backend.db.models import AnalysisResult, ProposalOutline


    # 1. 분석 결과 로드
    analysis = (await db.execute(
        select(AnalysisResult).where(
            AnalysisResult.notice_id == notice_id,
            AnalysisResult.analysis_status == "completed",
        )
    )).scalar_one_or_none()
    if not analysis:
        progress_store.emit(notice_id, "완료된 분석 결과가 없습니다", level="error")
        raise ValueError("완료된 분석 결과가 없습니다")

    project_type = analysis.project_type
    if not is_supported(project_type):
        # Gemini가 본문 보고 "기타" 등으로 분류했더라도 수집기가 공고명으로 ISP/ISMP를 판별했으면 그걸 폴백으로 사용.
        from backend.db.crud import get_notice_by_id
        notice = await get_notice_by_id(db, notice_id)
        fallback = notice.isp_ismp_type if notice else None
        if is_supported(fallback):
            progress_store.emit(
                notice_id,
                f"분석 결과 project_type={project_type or '미정'} 미지원 → 공고명 분류({fallback})로 폴백"
            )
            project_type = fallback
        else:
            labels = ", ".join(cfg["label"] for cfg in OUTLINE_TYPES.values())
            msg = (
                f"{get_label(project_type) or project_type or '미정'} 유형은 "
                f"목차 생성을 지원하지 않습니다 (지원: {labels})"
            )
            progress_store.emit(notice_id, msg, level="error")
            raise UnsupportedProjectTypeError(msg)

    progress_store.emit(notice_id, f"목차 생성 준비 ({get_label(project_type)})")

    raw = analysis.raw_analysis
    if isinstance(raw, str):
        analysis_data = json.loads(raw)
    elif isinstance(raw, dict):
        analysis_data = raw
    else:
        analysis_data = {}

    progress_store.emit(notice_id, "분석 결과 로드 완료")
    prompt = await build_outline_prompt_dynamic(analysis_data, project_type, db)

    # 2. LLM 호출 — provider 추상화 (Gemini/Claude/OpenAI).
    outline_system = await get_prompt("outline.system", db)
    llm_cfg = await get_active_config(db)
    # outline은 분석 결과 + 가이드라인 + 샘플이 동봉돼 input이 크므로 분석보다 살짝 높은 temperature.
    outline_temperature = max(0.0, min(2.0, llm_cfg.temperature + 0.1))
    request = LLMRequest(
        system=outline_system,
        user_text=prompt,
        model=llm_cfg.model,
        temperature=outline_temperature,
        response_json=True,
    )

    # 모델별 응답 타임아웃 (초). preview 모델이 큰 input에서 무응답으로 hang하는 케이스 방어.
    OUTLINE_LLM_TIMEOUT = 180
    try:
        response = await call_with_fallback(
            llm_cfg,
            request,
            timeout=OUTLINE_LLM_TIMEOUT,
            on_progress=lambda msg, lvl: progress_store.emit(notice_id, msg, level=lvl),
        )
    except LLMError as e:
        raise RuntimeError(f"제안목차 LLM 호출 실패: {e}") from e

    actual_model = response.model_used
    progress_store.emit(notice_id, f"응답 JSON 파싱 (model={actual_model})")
    outline_data = _parse_json_response(response.text)

    # 2.5. RFP 원문 (요구사항 섹션 verbatim) 추출 — 별도 LLM 호출 + 코드 슬라이스.
    rfp_raw_text: str | None = None
    try:
        from backend.db.models import Attachment
        from backend.services.rfp_extract import (
            build_rfp_text_from_attachments,
            extract_requirements_section,
        )

        rfp_atts = (await db.execute(
            select(Attachment).where(
                Attachment.notice_id == notice_id,
                Attachment.is_rfp.is_(True),
            ).order_by(Attachment.file_name)
        )).scalars().all()

        if rfp_atts:
            progress_store.emit(notice_id, f"RFP 원문 수집 ({len(rfp_atts)}건)")
            combined = await build_rfp_text_from_attachments(rfp_atts, db, notice_id)
            if combined:
                progress_store.emit(notice_id, "요구사항 섹션 앵커 추출")
                rfp_raw_text = await extract_requirements_section(combined, db, notice_id)
            else:
                progress_store.emit(notice_id, "RFP 첨부 변환 결과가 비어 원문 생략", level="warning")
        else:
            progress_store.emit(notice_id, "is_rfp 첨부 없음 — 원문 생략", level="warning")
    except Exception as e:  # noqa: BLE001
        logger.warning(f"RFP 원문 추출 실패 (notice_id={notice_id}): {e}")
        progress_store.emit(notice_id, f"RFP 원문 추출 실패: {str(e)[:80]}", level="warning")

    # 3. DB UPSERT — notice_id 기준 활성 outline 1건
    existing = (await db.execute(
        select(ProposalOutline).where(
            ProposalOutline.notice_id == notice_id,
            ProposalOutline.is_active.is_(True),
        )
    )).scalar_one_or_none()
    guideline_base = "ISMP" if project_type == "ISMP" else "MOIS_ISP"
    if existing:
        existing.sections = outline_data
        existing.model_used = actual_model
        existing.guideline_base = guideline_base
        existing.generated_at = datetime.now()
        existing.rfp_raw_text = rfp_raw_text
    else:
        db.add(ProposalOutline(
            notice_id=notice_id,
            sections=outline_data,
            guideline_base=guideline_base,
            model_used=actual_model,
            generated_at=datetime.now(),
            is_active=True,
            rfp_raw_text=rfp_raw_text,
        ))
    await db.commit()

    progress_store.emit(notice_id, "DB 저장 완료", level="success")
    logger.info(f"제안목차 생성 완료: notice_id={notice_id}, type={project_type}")
    # 엑셀 빌더가 원문 시트를 만들 수 있도록 rfp_raw_text를 outline_data에 동봉해 반환.
    outline_data = {**outline_data, "rfp_raw_text": rfp_raw_text}
    return outline_data


# ---------------------------------------------------------------------------
# 엑셀 빌더
# ---------------------------------------------------------------------------


_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
# 샘플 분석 기반 컬러 팔레트 (Malgun Gothic 9pt 기반)
_FILL_HEADER_GRAY = PatternFill(start_color="D8D8D8", end_color="D8D8D8", fill_type="solid")  # 평가배점 셀
_FILL_HEADER_LIGHT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Main 라벨
_FILL_TOP_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")    # 대분류 행
_FILL_HEADER_BLUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")   # Main 강조

_FONT_TITLE_14 = Font(name="Malgun Gothic", bold=True, size=14)
_FONT_HEADER_9_BOLD = Font(name="Malgun Gothic", bold=True, size=9)
_FONT_HEADER_12_BOLD = Font(name="Malgun Gothic", bold=True, size=12)
_FONT_HEADER_10_BOLD = Font(name="Malgun Gothic", bold=True, size=10)
_FONT_BODY_9 = Font(name="Malgun Gothic", size=9)
_FONT_BODY_10 = Font(name="Malgun Gothic", size=10)
_FONT_BODY_9_BOLD = Font(name="Malgun Gothic", size=9, bold=True)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_CENTER_NO_WRAP = Alignment(horizontal="center", vertical="center")


def _apply_border_range(ws, cell_range: str) -> None:
    """병합 셀 등 범위 전체에 테두리 적용."""
    for row in ws[cell_range]:
        for c in row:
            c.border = _BORDER


def build_outline_xlsx(outline_data: dict) -> bytes:
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Main"
    _build_main_sheet(ws_main, outline_data.get("main") or {})

    ws_outline = wb.create_sheet("제안 목차")
    _build_outline_sheet(ws_outline, outline_data)

    ws_req = wb.create_sheet("RFP 요구사항")
    _build_requirements_sheet(ws_req, outline_data.get("requirements") or [])

    raw_text = outline_data.get("rfp_raw_text") or ""
    if raw_text.strip():
        ws_raw = wb.create_sheet("RFP 원문")
        _build_raw_text_sheet(ws_raw, raw_text)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# RFP 원문 시트 — 단일 컬럼 A, 문단 단위(\n\n) split, wrap_text, 파일/섹션 헤더 강조
# ---------------------------------------------------------------------------


def _build_raw_text_sheet(ws, raw_text: str) -> None:
    ws.column_dimensions["A"].width = 100

    title = ws.cell(row=1, column=1, value="RFP 원문 (요구사항 섹션 verbatim)")
    title.font = _FONT_TITLE_14
    title.alignment = _CENTER
    title.fill = _FILL_HEADER_LIGHT
    title.border = _BORDER
    ws.row_dimensions[1].height = 30

    row = 2
    # \n\n 단위로 split — 빈 단락 제거
    paragraphs = [p.rstrip() for p in raw_text.split("\n\n") if p.strip()]
    for para in paragraphs:
        cell = ws.cell(row=row, column=1, value=para)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # 파일 구분자 / 섹션 헤더 강조 (`---` 또는 `#` 로 시작)
        if para.startswith("---") or para.startswith("# "):
            cell.font = _FONT_HEADER_10_BOLD
            cell.fill = _FILL_HEADER_LIGHT
        else:
            cell.font = _FONT_BODY_10
        cell.border = _BORDER
        row += 1


# ---------------------------------------------------------------------------
# Main 시트 — 라벨(A) + 값(B:H 병합), 사업명 14pt, 비고 큰 행
# ---------------------------------------------------------------------------
def _build_main_sheet(ws, main: dict) -> None:
    rows = [
        ("사업명", main.get("project_name", "")),
        ("고객명", main.get("client", "")),
        ("사업기간", main.get("duration", "")),
        ("용역금액", main.get("amount", "")),
        ("제출일", main.get("submit_date", "")),
        ("제출장소", main.get("submit_place", "")),
        ("비고 / 특이사항", main.get("notes", "")),
    ]
    for i, (label, value) in enumerate(rows, 1):
        # 라벨 (A 컬럼)
        c1 = ws.cell(row=i, column=1, value=label)
        c1.font = _FONT_HEADER_10_BOLD
        c1.fill = _FILL_HEADER_LIGHT
        c1.alignment = _CENTER
        # 값 (B:H 병합) — 사업명만 14pt
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
        c2 = ws.cell(row=i, column=2, value=str(value) if value is not None else "")
        c2.font = _FONT_TITLE_14 if label == "사업명" else _FONT_BODY_10
        c2.alignment = _LEFT
        # 행 전체 테두리
        _apply_border_range(ws, f"A{i}:H{i}")

    # 행 높이 (샘플 패턴)
    row_heights = {1: 34.5, 2: 22.5, 3: 22.5, 4: 22.5, 5: 22.5, 6: 22.5, 7: 110.0}
    for r, h in row_heights.items():
        ws.row_dimensions[r].height = h

    # 컬럼 폭
    ws.column_dimensions["A"].width = 19.5
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16

    # 이슈/공지 영역 (9행~)
    c_title = ws.cell(row=9, column=1, value="이슈 및 공지")
    c_title.font = _FONT_HEADER_12_BOLD
    c_title.alignment = _CENTER
    ws.merge_cells("A9:B9")
    issue_headers = ["no.", "date", "name", "message"]
    for col_off, h in enumerate(issue_headers):
        c = ws.cell(row=9, column=3 + col_off, value=h)
        c.font = _FONT_HEADER_10_BOLD
        c.fill = _FILL_HEADER_LIGHT
        c.alignment = _CENTER
        c.border = _BORDER
    for n in range(1, 12):
        c = ws.cell(row=9 + n, column=3, value=n)
        c.alignment = _CENTER
        c.border = _BORDER
        for col in (4, 5, 6):
            ws.cell(row=9 + n, column=col).border = _BORDER


# ---------------------------------------------------------------------------
# 제안 목차 시트 — 사업명 상단 병합, 헤더 2줄, 카테고리 셀 병합
# ---------------------------------------------------------------------------
_OUTLINE_COLS = 10  # A 평가배점 | B 관련ID | C 구분 | D L1 | E L2 | F L3 | G 담당자 | H 페이지 | I 완료여부 | J 비고
_OUTLINE_COL_WIDTHS = [21.9, 14.3, 14.1, 34.3, 36.7, 57.4, 13.9, 8.9, 7.1, 48.9]


def _style_row(ws, row_num: int, *, fill=None) -> None:
    """행 전체에 9pt 폰트 + 테두리 + (선택)배경 적용."""
    for col in range(1, _OUTLINE_COLS + 1):
        c = ws.cell(row=row_num, column=col)
        if c.font.name != "Malgun Gothic" or c.font.bold:
            pass  # 이미 설정된 셀(특수 셀)은 유지
        else:
            c.font = _FONT_BODY_9
        c.border = _BORDER
        if fill is not None:
            c.fill = fill


def _build_outline_sheet(ws, data: dict) -> None:
    main = data.get("main") or {}
    outline = data.get("outline") or []

    # 1행: 사업명 14pt center (A1:J1 병합)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_OUTLINE_COLS)
    c1 = ws.cell(row=1, column=1, value=main.get("project_name", ""))
    c1.font = _FONT_TITLE_14
    c1.alignment = _CENTER

    # 2행: 합계 행 (H2에 SUM 수식) — 샘플 패턴 따름
    c_h2 = ws.cell(row=2, column=8, value="합계")
    c_h2.font = _FONT_HEADER_9_BOLD
    c_h2.alignment = _CENTER
    c_h2.border = _BORDER

    # 헤더 (3~4행) — 2줄 구조, fill 없음 (샘플과 동일하게 흰색 헤더)
    ws.merge_cells("A3:A4")
    ws.merge_cells("B3:B4")
    ws.merge_cells("C3:F3")
    ws.merge_cells("G3:G4")
    ws.merge_cells("H3:H4")
    ws.merge_cells("I3:I4")
    ws.merge_cells("J3:J4")

    top_headers = {
        "A3": "평가항목 배점\n(제안요약서/발표자료에 필수)",
        "B3": "관련\n제안요청 ID",
        "C3": "목차",
        "G3": "담당자",
        "H3": "페이지수",
        "I3": "완료여부",
        "J3": "비고",
    }
    for addr, text in top_headers.items():
        c = ws[addr]
        c.value = text
        c.font = _FONT_HEADER_9_BOLD
        c.alignment = _CENTER
    # 4행 (C~F): 구분 / Level 1 / Level 2 / Level 3
    sub_headers = ["구분", "Level 1", "Level 2", "Level 3"]
    for i, h in enumerate(sub_headers):
        c = ws.cell(row=4, column=3 + i, value=h)
        c.font = _FONT_HEADER_9_BOLD
        c.alignment = _CENTER

    _apply_border_range(ws, f"A3:{get_column_letter(_OUTLINE_COLS)}4")

    # 본문 작성
    row_num = 5
    top_rows: list[int] = []
    for top in outline:
        top_rows.append(row_num)
        row_num = _write_section(ws, top, row_num)

    # H2 합계 — 대분류 행들의 페이지수 합산
    if top_rows:
        formula = "=" + "+".join(f"H{r}" for r in top_rows)
        ws.cell(row=2, column=8).value = formula

    # 컬럼 폭 (샘플 정확 수치)
    for i, w in enumerate(_OUTLINE_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # 행 높이 — 샘플은 1행 20.25, 헤더 12.75, 본문 12.75 통일
    ws.row_dimensions[1].height = 20.25
    ws.row_dimensions[2].height = 12.75
    ws.row_dimensions[3].height = 27
    ws.row_dimensions[4].height = 14.5


def _write_section(ws, top: dict, row_num: int) -> int:
    """대분류(L0 — Ⅰ., Ⅱ. 등) 행. 노란색(FFFF00) 배경."""
    code = top.get("code", "") or ""
    title = top.get("title", "") or ""
    eval_mapping = top.get("eval_mapping") or ""
    page_count = top.get("page_count")

    # A 평가배점 (있으면 회색 강조)
    c_eval = ws.cell(row=row_num, column=1, value=eval_mapping)
    c_eval.alignment = _CENTER
    c_eval.font = _FONT_BODY_9_BOLD if eval_mapping else _FONT_BODY_9
    if eval_mapping:
        c_eval.fill = _FILL_HEADER_GRAY
    # B 관련ID
    ws.cell(row=row_num, column=2, value="").alignment = _CENTER
    # C 구분 (대분류) — "Ⅰ. 일반현황" 형태
    division = f"{code}. {title}".strip(". ")
    c_div = ws.cell(row=row_num, column=3, value=division)
    c_div.font = _FONT_BODY_9_BOLD
    c_div.alignment = _CENTER
    c_div.fill = _FILL_TOP_YELLOW
    # D/E/F는 비워두되 노란 배경만
    for col in (4, 5, 6):
        cc = ws.cell(row=row_num, column=col)
        cc.fill = _FILL_TOP_YELLOW
    # G 담당자
    ws.cell(row=row_num, column=7, value="").alignment = _CENTER
    # H 페이지수 (노란 배경)
    c_pg = ws.cell(row=row_num, column=8, value=page_count)
    c_pg.alignment = _CENTER
    c_pg.fill = _FILL_TOP_YELLOW
    c_pg.font = _FONT_BODY_9_BOLD
    # I 완료여부
    ws.cell(row=row_num, column=9, value="").alignment = _CENTER
    # J 비고
    ws.cell(row=row_num, column=10, value="").alignment = _LEFT

    # 행 전체 테두리 + 미설정 셀에 폰트 적용
    for col in range(1, _OUTLINE_COLS + 1):
        c = ws.cell(row=row_num, column=col)
        if not c.font.bold:
            c.font = _FONT_BODY_9
        c.border = _BORDER

    row_num += 1
    return _write_children(ws, top.get("children") or [], row_num)


def _write_children(ws, nodes: list, row_num: int) -> int:
    """L1~L3 children. 일반 배경 (배경색 없음). 평가배점 셀만 회색."""
    for node in nodes:
        level = max(1, min(3, int(node.get("level", 1))))
        number = node.get("number", "") or ""
        title = node.get("title", "") or ""
        text = f"{number} {title}".strip()
        target_col = 3 + level  # L1=D(4), L2=E(5), L3=F(6)
        eval_mapping = node.get("eval_mapping") or ""

        # A 평가배점 (있으면 회색)
        c_eval = ws.cell(row=row_num, column=1, value=eval_mapping)
        c_eval.alignment = _CENTER
        c_eval.font = _FONT_BODY_9_BOLD if eval_mapping else _FONT_BODY_9
        if eval_mapping:
            c_eval.fill = _FILL_HEADER_GRAY
        # B 관련ID
        c_rel = ws.cell(row=row_num, column=2, value=node.get("related_req_id") or "")
        c_rel.alignment = _CENTER
        c_rel.font = _FONT_BODY_9
        # C~F (구분/L1/L2/L3) — target_col에만 값
        for col in (3, 4, 5, 6):
            cell = ws.cell(row=row_num, column=col, value="")
            cell.font = _FONT_BODY_9
            if col == target_col:
                cell.value = text
                cell.alignment = _LEFT
        # G 담당자
        cg = ws.cell(row=row_num, column=7, value=node.get("assignee", "") or "")
        cg.alignment = _CENTER
        cg.font = _FONT_BODY_9
        # H 페이지수
        ch = ws.cell(row=row_num, column=8, value=node.get("page_count"))
        ch.alignment = _CENTER
        ch.font = _FONT_BODY_9
        # I 완료여부
        ci = ws.cell(row=row_num, column=9, value="")
        ci.alignment = _CENTER
        ci.font = _FONT_BODY_9
        # J 비고
        cj = ws.cell(row=row_num, column=10, value=node.get("note", "") or "")
        cj.alignment = _LEFT
        cj.font = _FONT_BODY_9

        # 테두리
        for col in range(1, _OUTLINE_COLS + 1):
            ws.cell(row=row_num, column=col).border = _BORDER

        row_num += 1
        children = node.get("children") or []
        if children:
            row_num = _write_children(ws, children, row_num)
    return row_num


# ---------------------------------------------------------------------------
# RFP 요구사항 시트 — A1:G1 제목, 회색 헤더, 분류 셀 병합, E열 매우 넓게
# ---------------------------------------------------------------------------
_REQ_COL_WIDTHS = [14, 10, 30, 30, 90, 36, 20]  # 분류 / CODE / 명칭 / 정의 / 상세 / 산출 / 비고


def _build_requirements_sheet(ws, requirements: list) -> None:
    # 1행: 제목 (A1:G1 병합)
    ws.merge_cells("A1:G1")
    c_title = ws.cell(row=1, column=1, value="요구사항 목록")
    c_title.font = _FONT_TITLE_14
    c_title.alignment = _CENTER
    c_title.fill = _FILL_HEADER_LIGHT
    _apply_border_range(ws, "A1:G1")

    # 2행: 헤더
    headers = ["분류", "CODE", "요구사항 명칭", "정의", "요구사항 상세설명", "산출정보", "비고"]
    for col, label in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = _FONT_HEADER_12_BOLD
        c.fill = _FILL_HEADER_LIGHT
        c.alignment = _CENTER
        c.border = _BORDER

    # 본문 — 분류 연속된 행 묶기 위해 마지막 분류 시작 행 추적
    last_category = None
    category_start_row = None
    for i, req in enumerate(requirements, start=3):
        category = req.get("category", "") or ""
        # 같은 카테고리 묶음 셀 병합 처리
        if category != last_category:
            if last_category is not None and category_start_row is not None and category_start_row < i - 1:
                ws.merge_cells(start_row=category_start_row, start_column=1, end_row=i - 1, end_column=1)
            last_category = category
            category_start_row = i
            ws.cell(row=i, column=1, value=category)
        # 나머지 컬럼
        cells = [
            req.get("code", ""),
            req.get("name", ""),
            req.get("definition", ""),
            req.get("detail", ""),
            req.get("output", ""),
            req.get("note", ""),
        ]
        for col_off, value in enumerate(cells, start=2):
            c = ws.cell(row=i, column=col_off, value=value)
            c.font = _FONT_BODY_10
            c.alignment = _LEFT if col_off >= 3 else _CENTER
            c.border = _BORDER
        # 분류 셀 폰트/정렬
        c_cat = ws.cell(row=i, column=1)
        c_cat.font = _FONT_BODY_10
        c_cat.alignment = _CENTER
        c_cat.border = _BORDER
        c_cat.fill = _FILL_HEADER_LIGHT

    # 마지막 카테고리 병합 처리
    if last_category is not None and category_start_row is not None:
        last_row = 2 + len(requirements)
        if category_start_row < last_row:
            ws.merge_cells(start_row=category_start_row, start_column=1, end_row=last_row, end_column=1)

    # 컬럼 폭
    for i, w in enumerate(_REQ_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # 행 높이
    ws.row_dimensions[1].height = 36.6
    ws.row_dimensions[2].height = 22
